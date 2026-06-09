#!/usr/bin/env python3
"""MINI-XL v2.0 — Test suite for the installable package."""

import sys
import json
import tempfile
from pathlib import Path

from mini_xl.utils import (
    horodatage, horodatage_lisible, nettoyer_nom,
    generer_nom_xlsx, est_extension_compatible,
    fichier_est_vide, resoudre_repertoire,
    chemin_sortie, configurer_logging,
)
from mini_xl.scanner import scanner_repertoire, FichierInfo, verifier_avertissement_taille
from mini_xl.analyseur import analyser_fichier, detecter_en_tetes, ResultatAnalyse
from mini_xl.generateur import generer_xlsx

tests_reussis: int = 0
tests_echoues: int = 0


def _ok(nom: str) -> None:
    global tests_reussis
    tests_reussis += 1
    print(f"  PASS  {nom}")


def _echec(nom: str, detail: str = "") -> None:
    global tests_echoues
    tests_echoues += 1
    msg = f"  FAIL  {nom}"
    if detail:
        msg += f" -- {detail}"
    print(msg)


def test_utils() -> None:
    print("\n[UTILS]")
    ts = horodatage()
    _ok("horodatage format") if len(ts) == 15 and "_" in ts else _echec("horodatage", ts)
    tl = horodatage_lisible()
    _ok("horodatage_lisible") if "-" in tl and ":" in tl else _echec("horodatage_lisible")
    _ok("nettoyer_nom spaces") if nettoyer_nom("rapport ventes") == "rapport_ventes" else _echec("nettoyer_nom")
    _ok("nettoyer_nom special") if nettoyer_nom("f@#i!!ch..ier") == "fichier" else _echec("nettoyer_nom special")
    nom = generer_nom_xlsx("rapport ventes.csv")
    _ok("generer_nom_xlsx") if nom.startswith("rapport_ventes_") and nom.endswith(".xlsx") else _echec("generer_nom_xlsx", nom)
    _ok("ext compatible csv") if est_extension_compatible(Path("t.csv")) else _echec("ext csv")
    _ok("ext non compatible pdf") if not est_extension_compatible(Path("t.pdf")) else _echec("ext pdf")


def _creer_fichiers_test(rep: Path) -> None:
    (rep / "test.csv").write_text("nom,age,ville\nJean,25,Paris\nPaul,30,Lyon\n", encoding="utf-8")
    (rep / "test_sc.csv").write_text("a;b;c\n1;2;3\n4;5;6\n", encoding="utf-8")
    (rep / "test.tsv").write_text("x\ty\tz\n1\t2\t3\n4\t5\t6\n", encoding="utf-8")
    (rep / "test.json").write_text('[{"nom":"A","val":10},{"nom":"B","val":20}]', encoding="utf-8")
    (rep / "test.txt").write_text("ref\tdesc\tprix\nA001\tClavier\t29.99\nA002\tSouris\t14.50\n", encoding="utf-8")


def test_scanner() -> None:
    print("\n[SCANNER]")
    with tempfile.TemporaryDirectory() as td:
        _creer_fichiers_test(Path(td))
        fichiers = scanner_repertoire(td)
        _ok(f"scan {len(fichiers)} files") if len(fichiers) == 5 else _echec("scan", str(len(fichiers)))
        exts = {f.extension for f in fichiers}
        _ok("all extensions") if exts == {".csv", ".tsv", ".json", ".txt"} else _echec("extensions", str(exts))


def test_analyseur() -> None:
    print("\n[ANALYZER]")
    with tempfile.TemporaryDirectory() as td:
        p = Path(td)
        _creer_fichiers_test(p)
        r = analyser_fichier(p / "test.csv")
        _ok("CSV type") if r.type_detecte == "csv" else _echec("CSV type", r.type_detecte)
        _ok("CSV headers") if r.en_tetes == ["nom", "age", "ville"] else _echec("CSV headers", str(r.en_tetes))
        _ok("CSV 2 rows") if len(r.donnees) == 2 else _echec("CSV rows", str(len(r.donnees)))
        r2 = analyser_fichier(p / "test_sc.csv")
        _ok("CSV SC type") if r2.type_detecte in ("csv_sc", "csv") else _echec("CSV SC", r2.type_detecte)
        r3 = analyser_fichier(p / "test.tsv")
        _ok("TSV type") if r3.type_detecte == "tsv" else _echec("TSV type", r3.type_detecte)
        r4 = analyser_fichier(p / "test.json")
        _ok("JSON type") if r4.type_detecte == "json" else _echec("JSON type", r4.type_detecte)
        _ok("JSON 2 rows") if len(r4.donnees) == 2 else _echec("JSON rows")
        r5 = analyser_fichier(p / "test.txt")
        _ok(f"TXT type ({r5.type_detecte})") if r5.type_detecte in ("txt", "tsv") else _echec("TXT type", r5.type_detecte)


def test_json_invalide() -> None:
    print("\n[INVALID JSON]")
    with tempfile.TemporaryDirectory() as td:
        (Path(td) / "bad.json").write_text('{"config":{"theme":"dark"}}', encoding="utf-8")
        try:
            analyser_fichier(Path(td) / "bad.json")
            _echec("non-array JSON should fail")
        except ValueError:
            _ok("non-array JSON rejected")


def test_fichier_vide() -> None:
    print("\n[EMPTY FILE]")
    with tempfile.TemporaryDirectory() as td:
        (Path(td) / "vide.csv").write_text("", encoding="utf-8")
        try:
            analyser_fichier(Path(td) / "vide.csv")
            _echec("empty file should fail")
        except ValueError:
            _ok("empty file rejected")


def test_en_tetes() -> None:
    print("\n[HEADER DETECTION]")
    _ok("headers detected") if detecter_en_tetes(["nom", "age"], [["Jean", "25"]]) else _echec("headers yes")
    _ok("no headers num") if not detecter_en_tetes(["100", "200"], [["300", "400"]]) else _echec("headers num")


def test_generateur() -> None:
    print("\n[EXCEL GENERATOR]")
    with tempfile.TemporaryDirectory() as td:
        sortie = Path(td) / "out.xlsx"
        generer_xlsx(["nom", "age"], [["Jean", "25"], ["Paul", "30"]], sortie, "test")
        _ok("file created") if sortie.exists() and sortie.stat().st_size > 0 else _echec("file created")
        from openpyxl import load_workbook
        wb = load_workbook(str(sortie))
        ws = wb.active
        _ok("sheet 'test'") if ws.title == "test" else _echec("sheet", ws.title)
        _ok("header nom") if ws.cell(1, 1).value == "nom" else _echec("header")
        _ok("data Jean") if ws.cell(2, 1).value == "Jean" else _echec("data")
        _ok("freeze A2") if ws.freeze_panes == "A2" else _echec("freeze", str(ws.freeze_panes))
        _ok("auto-filter") if ws.auto_filter.ref else _echec("filter")
        _ok("age=25 (int)") if isinstance(ws.cell(2, 2).value, (int, float)) else _echec("typing")
        wb.close()


def test_package_import() -> None:
    print("\n[PACKAGE IMPORT]")
    try:
        import mini_xl
        _ok(f"mini_xl v{mini_xl.__version__}")
    except ImportError as e:
        _echec("mini_xl import", str(e))

    try:
        from mini_xl.main import main
        _ok("mini_xl.main:main callable")
    except ImportError as e:
        _echec("entry point", str(e))


def main() -> int:
    print("=" * 50)
    print("  MINI-XL v2.0 — Test Suite (Package)")
    print("=" * 50)
    test_utils()
    test_scanner()
    test_analyseur()
    test_json_invalide()
    test_fichier_vide()
    test_en_tetes()
    test_generateur()
    test_package_import()
    total = tests_reussis + tests_echoues
    print(f"\n  Results: {tests_reussis}/{total} passed, {tests_echoues} failed\n")
    return 1 if tests_echoues > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
