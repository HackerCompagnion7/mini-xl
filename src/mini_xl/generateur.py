"""MINI-XL v2.0 - Professional Excel generator."""

import logging
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

POLICE_EN_TETE: str = "Calibri"
TAILLE_EN_TETE: int = 11
TAILLE_DONNEES: int = 10
COULEUR_FOND: str = "4472C4"
COULEUR_TEXTE: str = "FFFFFF"
BORDURE = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def generer_xlsx(
    en_tetes: list[str], donnees: list[list[str]],
    chemin_sortie: Path, nom_feuille: str = "Data",
    logger: Optional[logging.Logger] = None,
) -> Path:
    """Generate a professionally formatted Excel file."""
    if not en_tetes or any(not h.strip() for h in en_tetes):
        raise ValueError("Invalid or empty headers")

    wb = Workbook()
    ws = wb.active
    ws.title = _nettoyer_nom_feuille(nom_feuille)

    _ecrire_en_tetes(ws, en_tetes)
    _ecrire_donnees(ws, donnees)
    _ajuster_largeurs(ws, en_tetes, donnees)
    _activer_filtre(ws, en_tetes)
    ws.freeze_panes = "A2"

    wb.save(str(chemin_sortie))
    _log(logger, f"Excel generated: {chemin_sortie} ({len(donnees)} rows)")
    return chemin_sortie


def _ecrire_en_tetes(ws: object, en_tetes: list[str]) -> None:
    """Write the header row with bold colored style."""
    police = Font(
        name=POLICE_EN_TETE, size=TAILLE_EN_TETE,
        bold=True, color=COULEUR_TEXTE,
    )
    fond = PatternFill(
        start_color=COULEUR_FOND, end_color=COULEUR_FOND,
        fill_type="solid",
    )
    align = Alignment(horizontal="center", vertical="center")

    for col, en_tete in enumerate(en_tetes, start=1):
        cell = ws.cell(row=1, column=col, value=en_tete)
        cell.font = police
        cell.fill = fond
        cell.alignment = align
        cell.border = BORDURE


def _ecrire_donnees(ws: object, donnees: list[list[str]]) -> None:
    """Write data rows with automatic type conversion."""
    police = Font(name=POLICE_EN_TETE, size=TAILLE_DONNEES)
    align = Alignment(vertical="center")

    for lig, ligne in enumerate(donnees, start=2):
        for col, valeur in enumerate(ligne, start=1):
            cell = ws.cell(row=lig, column=col, value=_convertir(valeur))
            cell.font = police
            cell.alignment = align
            cell.border = BORDURE


def _convertir(valeur: str) -> object:
    """Convert a string to its natural type (int, float or str)."""
    valeur = valeur.strip()
    if not valeur:
        return ""
    try:
        return int(valeur)
    except ValueError:
        pass
    try:
        return float(valeur.replace(",", "."))
    except ValueError:
        pass
    return valeur


def _ajuster_largeurs(
    ws: object, en_tetes: list[str], donnees: list[list[str]],
) -> None:
    """Adjust column widths to content."""
    for col in range(1, len(en_tetes) + 1):
        lettre = get_column_letter(col)
        largeur = _largeur_colonne(col, en_tetes, donnees)
        ws.column_dimensions[lettre].width = largeur


def _largeur_colonne(
    col: int, en_tetes: list[str], donnees: list[list[str]],
) -> float:
    """Calculate optimal width for a column."""
    max_len = len(str(en_tetes[col - 1]))
    for ligne in donnees:
        if col <= len(ligne):
            l = len(str(ligne[col - 1]))
            if l > max_len:
                max_len = l
    return min(max_len + 3, 50)


def _activer_filtre(ws: object, en_tetes: list[str]) -> None:
    """Enable auto-filter on the header row."""
    fin = get_column_letter(len(en_tetes))
    ws.auto_filter.ref = f"A1:{fin}1"


def _nettoyer_nom_feuille(nom: str) -> str:
    """Clean sheet name (max 31 chars, no forbidden chars)."""
    for car in r"[]:*?/\\.":
        nom = nom.replace(car, "_")
    return nom[:31]


def _log(logger: Optional[logging.Logger], msg: str) -> None:
    """Log if logger is available."""
    if logger:
        logger.info(msg)
